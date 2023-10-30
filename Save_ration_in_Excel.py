import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.worksheet.page import PageMargins
# Подключаем библиотеку для работы с окнами
# import tkinter as tk
from tkinter import filedialog

def save_file(data_ration, data, data_input, text_title):
    """Функция открытия диалогового окна для ввода имени файла в формате Excel для сохранения рациона"""
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    # Выбранное имя файла будет храниться в переменной file_path
    if file_path:

        # Создание новой рабочей книги
        workbook = openpyxl.Workbook()

        # Выбор активного листа
        sheet = workbook.active
        sheet.title = 'Рацион'

        # Настройка параметров страницы
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4  # Размер бумаги (A4)
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE  # Ориентация (горизонтальная)
        sheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)  # Поля (в дюймах)

        # Задание ширины столбцов
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10

        # Создание объекта Border для определения границ
        border = Border(
            left=Side(border_style="thin", color="000000"),  # Левая граница
            right=Side(border_style="thin", color="000000"),  # Правая граница
            top=Side(border_style="thin", color="000000"),  # Верхняя граница
            bottom=Side(border_style="thin", color="000000"),  # Нижняя граница
        )

        # Объединение ячеек
        sheet.merge_cells("A1:C1")  # Объединение ячеек A1:C1
        sheet['A1'] = text_title
        sheet.merge_cells("A3:F3")  # Объединение ячеек A3:F3
        # Установка высоты строки
        sheet.row_dimensions[3].height = 30
        sheet['A3'] = data_input
        # Установка переноса текста
        sheet['A3'].alignment = openpyxl.styles.Alignment(wrapText=True)

        sheet['A5'] = "Состав рациона:"

        for row_data in data_ration:
            sheet.append(row_data)

        # Применение границ к диапазону ячеек, например, к таблице данных
        for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border


        sheet['A' + str(sheet.max_row + 2)] = "Питательность рациона:"
        _m = sheet.max_row + 1

        for row_data in data:
            sheet.append(row_data)

        # Применение границ к диапазону ячеек, например, к таблице данных
        for row in sheet.iter_rows(min_row=_m, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border

        # Сохранение рабочей книги в файл Excel file_path
        workbook.save(file_path)
